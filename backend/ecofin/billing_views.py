"""
Eco-Fin Billing Views
API endpoints pentru facturare SmartBill.
"""
import os
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.core.mail import EmailMessage
from django.db import transaction
from django.db.models import Sum, Q
from django.http import HttpResponse, FileResponse
from django.utils import timezone

from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from iss.models import Client
from .models import (
    BillingInvoice, 
    BillingInvoiceLine, 
    BillingSyncLog,
    BillingEmailLog,
    EcoFinProcessedRecord
)
from .serializers import (
    BillingInvoiceSerializer,
    BillingInvoiceListSerializer,
    BillingInvoiceLineSerializer,
    BillingSyncLogSerializer,
    BillingEmailLogSerializer,
    IssueInvoiceRequestSerializer,
    InvoicePreviewSerializer,
    SendEmailRequestSerializer,
    BillingReportFilterSerializer
)
from .smartbill_client import (
    SmartBillClient, 
    SmartBillError, 
    get_smartbill_client,
    is_smartbill_configured
)


# Mapping pentru numele lunilor în română
MONTH_NAMES_RO = {
    1: 'IANUARIE', 2: 'FEBRUARIE', 3: 'MARTIE', 4: 'APRILIE',
    5: 'MAI', 6: 'IUNIE', 7: 'IULIE', 8: 'AUGUST',
    9: 'SEPTEMBRIE', 10: 'OCTOMBRIE', 11: 'NOIEMBRIE', 12: 'DECEMBRIE'
}


class IsManagementOrAdmin:
    """
    Permisiune: doar utilizatorii cu rol Management sau Admin.
    """
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        if request.user.is_superuser:
            return True
        # Verifică rolul utilizatorului
        role = getattr(request.user, 'role', None)
        return role in ['management', 'admin', 'Management', 'Admin']


class BillingInvoiceViewSet(viewsets.ModelViewSet):
    """
    ViewSet pentru gestionarea facturilor.
    
    Endpoints:
    - GET /api/eco-fin/billing/invoices/ - Lista facturi
    - GET /api/eco-fin/billing/invoices/{id}/ - Detalii factură
    - POST /api/eco-fin/billing/invoices/preview/ - Preview factură
    - POST /api/eco-fin/billing/invoices/issue/ - Emite factură
    - GET /api/eco-fin/billing/invoices/{id}/pdf/ - Descarcă PDF
    - POST /api/eco-fin/billing/invoices/{id}/send-email/ - Trimite email
    """
    queryset = BillingInvoice.objects.all()
    serializer_class = BillingInvoiceSerializer
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return BillingInvoiceListSerializer
        return BillingInvoiceSerializer
    
    def get_queryset(self):
        queryset = BillingInvoice.objects.select_related('client', 'created_by')
        
        # Filtre
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        client_id = self.request.query_params.get('client_id')
        payment_status = self.request.query_params.get('payment_status')
        last_months = self.request.query_params.get('last_months')
        
        if year:
            queryset = queryset.filter(year=int(year))
        if month:
            queryset = queryset.filter(month=int(month))
        if client_id:
            queryset = queryset.filter(client_id=int(client_id))
        if payment_status and payment_status != 'all':
            queryset = queryset.filter(payment_status=payment_status)
        
        # Filtru ultimele N luni
        if last_months:
            today = datetime.now()
            months_back = int(last_months)
            date_filters = Q()
            for i in range(months_back):
                target_date = today - timedelta(days=30 * i)
                date_filters |= Q(year=target_date.year, month=target_date.month)
            queryset = queryset.filter(date_filters)
        
        return queryset.order_by('-year', '-month', '-issue_date')
    
    @action(detail=False, methods=['get'], url_path='check-config')
    def check_config(self, request):
        """Verifică dacă SmartBill este configurat."""
        configured = is_smartbill_configured()
        
        if configured:
            client = get_smartbill_client()
            if client:
                test_result = client.test_connection()
                return Response({
                    'configured': True,
                    'connection_test': test_result
                })
        
        return Response({
            'configured': False,
            'message': 'SmartBill credentials nu sunt configurate. '
                      'Setați SMARTBILL_USERNAME, SMARTBILL_TOKEN, SMARTBILL_COMPANY_CIF în environment.'
        })
    
    @action(detail=False, methods=['post'], url_path='preview')
    def preview(self, request):
        """
        Generează preview pentru factură.
        Returnează datele calculate + avertismente.
        """
        client_id = request.data.get('client_id')
        year = request.data.get('year')
        month = request.data.get('month')
        
        if not all([client_id, year, month]):
            return Response(
                {'detail': 'client_id, year și month sunt obligatorii'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            client = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            return Response(
                {'detail': 'Clientul nu a fost găsit'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Obține orele din EcoFinProcessedRecord
        records = EcoFinProcessedRecord.objects.filter(
            client=client,
            year=year,
            month=month
        )
        
        total_hours = records.aggregate(Sum('ore_lucrate'))['ore_lucrate__sum'] or Decimal('0')
        hourly_rate = getattr(client, 'tarif_orar', Decimal('0')) or Decimal('0')
        
        # Calculează valoarea
        subtotal = total_hours * hourly_rate
        vat_rate = Decimal('21')
        vat_total = subtotal * (vat_rate / 100)
        total = subtotal + vat_total
        
        # Verifică facturi existente pentru aceeași lună
        existing_invoices = BillingInvoice.objects.filter(
            client=client,
            year=year,
            month=month,
            status=BillingInvoice.InvoiceStatus.ISSUED
        )
        
        already_billed = existing_invoices.aggregate(Sum('subtotal'))['subtotal__sum'] or Decimal('0')
        
        warnings = []
        
        if existing_invoices.exists():
            if already_billed >= subtotal:
                warnings.append(
                    f'Există deja factură/facturi pentru {MONTH_NAMES_RO[int(month)]} {year} '
                    f'cu valoare totală {already_billed} RON (≥ {subtotal} RON calculat). '
                    f'Doriți să facturați alte servicii?'
                )
            else:
                difference = subtotal - already_billed
                warnings.append(
                    f'Există factură/facturi pentru {MONTH_NAMES_RO[int(month)]} {year} '
                    f'cu valoare {already_billed} RON. Diferență de facturat: {difference} RON.'
                )
        
        if total_hours == 0:
            warnings.append('Nu există ore înregistrate pentru această perioadă!')
        
        if hourly_rate == 0:
            warnings.append('Tariful orar pentru client este 0!')
        
        # Construiește liniile
        month_name = MONTH_NAMES_RO.get(int(month), str(month))
        lines = [{
            'description': f'PRESTARI SERVICII {month_name} {year}',
            'quantity': 1,
            'unit_price': float(subtotal),
            'vat_rate': float(vat_rate),
            'line_total': float(subtotal),
            'line_vat': float(vat_total)
        }]
        
        existing_list = [{
            'id': inv.id,
            'series_number': inv.invoice_number_display,
            'subtotal': float(inv.subtotal),
            'total': float(inv.total),
            'issue_date': inv.issue_date.isoformat() if inv.issue_date else None
        } for inv in existing_invoices]
        
        preview_data = {
            'client_id': client.id,
            'client_name': client.denumire,
            'year': int(year),
            'month': int(month),
            'month_name': month_name,
            'total_hours': float(total_hours),
            'hourly_rate': float(hourly_rate),
            'lines': lines,
            'subtotal': float(subtotal),
            'vat_rate': float(vat_rate),
            'vat_total': float(vat_total),
            'total': float(total),
            'existing_invoices': existing_list,
            'already_billed_amount': float(already_billed),
            'warnings': warnings
        }
        
        return Response(preview_data)
    
    @action(detail=False, methods=['post'], url_path='issue')
    def issue_invoice(self, request):
        """
        Emite o factură în SmartBill.
        
        Payload:
        {
            "client_id": 1,
            "year": 2025,
            "month": 11,
            "confirm_hours_agreed": true,
            "mode": "standard" | "difference" | "extra_services",
            "extra_lines": [...],  // opțional
            "issue_difference": false  // opțional
        }
        """
        serializer = IssueInvoiceRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        
        # Verifică confirmarea orelor
        if not data['confirm_hours_agreed']:
            return Response(
                {'detail': 'Trebuie să confirmați că numărul de ore este agreat cu clientul.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verifică configurarea SmartBill
        smartbill = get_smartbill_client()
        if not smartbill:
            return Response(
                {'detail': 'SmartBill nu este configurat. Contactați administratorul.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        
        try:
            client = Client.objects.get(id=data['client_id'])
        except Client.DoesNotExist:
            return Response(
                {'detail': 'Clientul nu a fost găsit'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        year = data['year']
        month = data['month']
        mode = data['mode']
        
        # Obține orele din EcoFinProcessedRecord
        records = EcoFinProcessedRecord.objects.filter(
            client=client, year=year, month=month
        )
        
        total_hours = records.aggregate(Sum('ore_lucrate'))['ore_lucrate__sum'] or Decimal('0')
        hourly_rate = getattr(client, 'tarif_orar', Decimal('0')) or Decimal('0')
        subtotal = total_hours * hourly_rate
        
        # Verifică facturi existente
        existing_invoices = BillingInvoice.objects.filter(
            client=client, year=year, month=month,
            status=BillingInvoice.InvoiceStatus.ISSUED
        )
        already_billed = existing_invoices.aggregate(Sum('subtotal'))['subtotal__sum'] or Decimal('0')
        
        # Construiește liniile în funcție de mod
        month_name = MONTH_NAMES_RO.get(month, str(month))
        lines = []
        vat_rate = Decimal('21')
        
        if mode == 'standard':
            # Factură standard pentru servicii complete
            lines.append({
                'name': f'PRESTARI SERVICII {month_name} {year}',
                'quantity': 1,
                'price': float(subtotal),
                'vatPercent': float(vat_rate),
                'um': 'buc'
            })
        
        elif mode == 'difference':
            # Factură pentru diferență
            if subtotal <= already_billed:
                return Response(
                    {'detail': 'Nu există diferență de facturat. Valoarea calculată nu depășește facturile existente.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            difference = subtotal - already_billed
            lines.append({
                'name': f'PRESTARI SERVICII {month_name} {year} – DIFERENȚĂ',
                'quantity': 1,
                'price': float(difference),
                'vatPercent': float(vat_rate),
                'um': 'buc'
            })
            subtotal = difference
        
        elif mode == 'extra_services':
            # Servicii suplimentare
            extra_lines = data.get('extra_lines', [])
            if not extra_lines:
                return Response(
                    {'detail': 'Pentru modul extra_services, trebuie să furnizați extra_lines.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            subtotal = Decimal('0')
            for extra in extra_lines:
                line_subtotal = Decimal(str(extra.get('quantity', 1))) * Decimal(str(extra.get('unit_price', 0)))
                lines.append({
                    'name': extra.get('description', 'Serviciu suplimentar'),
                    'quantity': float(extra.get('quantity', 1)),
                    'price': float(extra.get('unit_price', 0)),
                    'vatPercent': float(extra.get('vat_rate', 21)),
                    'um': extra.get('um', 'buc')
                })
                subtotal += line_subtotal
        
        # Calculează totalurile
        vat_total = subtotal * (vat_rate / 100)
        total = subtotal + vat_total
        
        # Pregătește datele clientului pentru SmartBill
        client_data = {
            'name': client.denumire,
            'cif': getattr(client, 'cif', ''),
            'address': getattr(client, 'adresa', ''),
            'city': getattr(client, 'oras', ''),
            'county': getattr(client, 'judet', ''),
            'country': 'România',
            'email': getattr(client, 'email', ''),
            'is_tax_payer': True
        }
        
        try:
            with transaction.atomic():
                # Emite factura în SmartBill
                result = smartbill.issue_invoice(
                    client_data=client_data,
                    lines=lines,
                    issue_date=datetime.now()
                )
                
                # Creează înregistrarea în BD
                invoice = BillingInvoice.objects.create(
                    client=client,
                    year=year,
                    month=month,
                    smartbill_series=result.get('series', ''),
                    smartbill_number=result.get('number', ''),
                    issue_date=datetime.now().date(),
                    subtotal=subtotal,
                    vat_total=vat_total,
                    total=total,
                    hours_billed=total_hours if mode == 'standard' else Decimal('0'),
                    hourly_rate=hourly_rate,
                    status=BillingInvoice.InvoiceStatus.ISSUED,
                    created_by=request.user
                )
                
                # Salvează liniile
                for i, line in enumerate(lines):
                    line_total = Decimal(str(line['quantity'])) * Decimal(str(line['price']))
                    line_vat = line_total * (Decimal(str(line['vatPercent'])) / 100)
                    BillingInvoiceLine.objects.create(
                        invoice=invoice,
                        description=line['name'],
                        quantity=Decimal(str(line['quantity'])),
                        unit_price=Decimal(str(line['price'])),
                        vat_rate=Decimal(str(line['vatPercent'])),
                        line_total=line_total,
                        line_vat=line_vat,
                        line_type='standard' if mode == 'standard' else 
                                 ('difference' if mode == 'difference' else 'extra')
                    )
                
                # Descarcă și salvează PDF-ul
                try:
                    pdf_content = smartbill.get_invoice_pdf(
                        series=result.get('series', ''),
                        number=result.get('number', '')
                    )
                    
                    # Salvează PDF-ul
                    pdf_dir = os.path.join(
                        settings.MEDIA_ROOT, 'invoices', 
                        str(client.id), str(year), str(month)
                    )
                    os.makedirs(pdf_dir, exist_ok=True)
                    
                    pdf_filename = f"{result.get('series', '')}{result.get('number', '')}.pdf"
                    pdf_path = os.path.join(pdf_dir, pdf_filename)
                    
                    with open(pdf_path, 'wb') as f:
                        f.write(pdf_content)
                    
                    invoice.pdf_path = os.path.join(
                        'invoices', str(client.id), str(year), str(month), pdf_filename
                    )
                    invoice.save()
                    
                except SmartBillError as e:
                    # PDF-ul nu a putut fi descărcat, dar factura a fost emisă
                    pass
                
                serializer = BillingInvoiceSerializer(invoice)
                return Response({
                    'success': True,
                    'message': f'Factura {invoice.invoice_number_display} a fost emisă cu succes.',
                    'invoice': serializer.data
                }, status=status.HTTP_201_CREATED)
                
        except SmartBillError as e:
            return Response({
                'success': False,
                'detail': f'Eroare SmartBill: {e.message}',
                'error_data': e.response_data
            }, status=status.HTTP_502_BAD_GATEWAY)
        except Exception as e:
            return Response({
                'success': False,
                'detail': f'Eroare la emiterea facturii: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'], url_path='pdf')
    def download_pdf(self, request, pk=None):
        """Descarcă PDF-ul facturii."""
        invoice = self.get_object()
        
        if not invoice.pdf_path:
            # Încearcă să descarce din SmartBill
            if invoice.smartbill_series and invoice.smartbill_number:
                smartbill = get_smartbill_client()
                if smartbill:
                    try:
                        pdf_content = smartbill.get_invoice_pdf(
                            series=invoice.smartbill_series,
                            number=invoice.smartbill_number
                        )
                        response = HttpResponse(pdf_content, content_type='application/pdf')
                        response['Content-Disposition'] = (
                            f'attachment; filename="{invoice.invoice_number_display}.pdf"'
                        )
                        return response
                    except SmartBillError:
                        pass
            
            return Response(
                {'detail': 'PDF-ul nu este disponibil'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Returnează PDF-ul salvat local
        pdf_full_path = os.path.join(settings.MEDIA_ROOT, invoice.pdf_path)
        if os.path.exists(pdf_full_path):
            return FileResponse(
                open(pdf_full_path, 'rb'),
                content_type='application/pdf',
                as_attachment=True,
                filename=f'{invoice.invoice_number_display}.pdf'
            )
        
        return Response(
            {'detail': 'Fișierul PDF nu a fost găsit'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    @action(detail=True, methods=['post'], url_path='send-email')
    def send_email(self, request, pk=None):
        """Trimite factura pe email."""
        invoice = self.get_object()
        
        serializer = SendEmailRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Obține email-ul destinatar
        email_to = serializer.validated_data.get('email_to')
        if not email_to:
            email_to = getattr(invoice.client, 'email', None)
        
        if not email_to:
            return Response(
                {'detail': 'Nu există adresă de email. Specificați email_to sau configurați email-ul clientului.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verifică dacă PDF-ul există
        if not invoice.pdf_path:
            return Response(
                {'detail': 'PDF-ul facturii nu este disponibil pentru trimitere.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        pdf_full_path = os.path.join(settings.MEDIA_ROOT, invoice.pdf_path)
        if not os.path.exists(pdf_full_path):
            return Response(
                {'detail': 'Fișierul PDF nu a fost găsit.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Construiește email-ul
        month_name = MONTH_NAMES_RO.get(invoice.month, str(invoice.month))
        subject = f'Factura {invoice.invoice_number_display} – {invoice.client.denumire} – {month_name}/{invoice.year}'
        
        body = f"""Bună ziua,

Vă trimitem atașat factura {invoice.invoice_number_display} pentru serviciile prestate în luna {month_name} {invoice.year}.

Detalii factură:
- Valoare fără TVA: {invoice.subtotal} RON
- TVA: {invoice.vat_total} RON
- Total: {invoice.total} RON

Cu respect,
International Staff Sourcing SRL
"""
        
        try:
            email = EmailMessage(
                subject=subject,
                body=body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email_to]
            )
            
            with open(pdf_full_path, 'rb') as f:
                email.attach(
                    f'{invoice.invoice_number_display}.pdf',
                    f.read(),
                    'application/pdf'
                )
            
            email.send()
            
            # Loghează trimiterea
            BillingEmailLog.objects.create(
                invoice=invoice,
                sent_by=request.user,
                sent_to=email_to,
                subject=subject,
                status='sent'
            )
            
            # Actualizează factura
            invoice.last_email_sent_at = timezone.now()
            invoice.email_sent_to = email_to
            invoice.email_sent_count += 1
            invoice.save()
            
            return Response({
                'success': True,
                'message': f'Email trimis cu succes la {email_to}'
            })
            
        except Exception as e:
            BillingEmailLog.objects.create(
                invoice=invoice,
                sent_by=request.user,
                sent_to=email_to,
                subject=subject,
                status='failed',
                error_message=str(e)
            )
            
            return Response({
                'success': False,
                'detail': f'Eroare la trimiterea email-ului: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class BillingSyncViewSet(viewsets.ViewSet):
    """
    ViewSet pentru sincronizarea plăților din SmartBill.
    """
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['post'], url_path='sync-payments')
    def sync_payments(self, request):
        """
        Sincronizează plățile din SmartBill (incremental).
        Folosește timestamp salvat pentru a cere doar modificările recente.
        """
        smartbill = get_smartbill_client()
        if not smartbill:
            return Response(
                {'detail': 'SmartBill nu este configurat.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        
        # Obține ultimul sync reușit
        last_sync = BillingSyncLog.objects.filter(
            status=BillingSyncLog.Status.SUCCESS
        ).order_by('-sync_finished_at').first()
        
        now = timezone.now()
        
        if last_sync and last_sync.sync_finished_at:
            from_date = last_sync.sync_finished_at
        else:
            # Prima sincronizare - ultimele 90 de zile
            from_date = now - timedelta(days=90)
        
        # Creează log de sync
        sync_log = BillingSyncLog.objects.create(
            requested_from_ts=from_date,
            requested_to_ts=now,
            user=request.user,
            status=BillingSyncLog.Status.IN_PROGRESS
        )
        
        try:
            # Obține plățile din SmartBill
            payments = smartbill.get_payments(
                from_date=from_date,
                to_date=now
            )
            
            invoices_updated = 0
            errors = []
            
            for payment in payments:
                try:
                    # Găsește factura locală
                    series = payment.get('invoiceSeries', '')
                    number = payment.get('invoiceNumber', '')
                    
                    invoice = BillingInvoice.objects.filter(
                        smartbill_series=series,
                        smartbill_number=number
                    ).first()
                    
                    if invoice:
                        # Actualizează statusul de plată
                        paid_amount = Decimal(str(payment.get('paidAmount', 0)))
                        invoice.paid_amount = paid_amount
                        invoice.last_payment_sync_at = now
                        invoice.save()  # save() actualizează automat due_amount și payment_status
                        invoices_updated += 1
                        
                except Exception as e:
                    errors.append(f"Eroare la procesarea plății: {str(e)}")
            
            # Actualizează log-ul
            sync_log.sync_finished_at = timezone.now()
            sync_log.status = BillingSyncLog.Status.SUCCESS
            sync_log.result_counts = {
                'payments_found': len(payments),
                'invoices_updated': invoices_updated,
                'errors_count': len(errors)
            }
            if errors:
                sync_log.error_message = '\n'.join(errors[:10])  # Primele 10 erori
            sync_log.save()
            
            return Response({
                'success': True,
                'sync_log_id': sync_log.id,
                'invoices_updated': invoices_updated,
                'payments_found': len(payments),
                'errors': errors[:5],
                'message': f'Sincronizare completă. {invoices_updated} facturi actualizate.'
            })
            
        except SmartBillError as e:
            sync_log.sync_finished_at = timezone.now()
            sync_log.status = BillingSyncLog.Status.FAILURE
            sync_log.error_message = str(e)
            sync_log.save()
            
            return Response({
                'success': False,
                'detail': f'Eroare SmartBill: {e.message}'
            }, status=status.HTTP_502_BAD_GATEWAY)
        
        except Exception as e:
            sync_log.sync_finished_at = timezone.now()
            sync_log.status = BillingSyncLog.Status.FAILURE
            sync_log.error_message = str(e)
            sync_log.save()
            
            return Response({
                'success': False,
                'detail': f'Eroare la sincronizare: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='sync-logs')
    def sync_logs(self, request):
        """Returnează istoricul sincronizărilor."""
        logs = BillingSyncLog.objects.order_by('-sync_started_at')[:50]
        serializer = BillingSyncLogSerializer(logs, many=True)
        return Response(serializer.data)


# ==========================================
# RAPOARTE FACTURARE
# ==========================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def billing_report_summary(request):
    """
    Raport sumar facturare cu filtre.
    
    GET /api/eco-fin/billing/reports/summary/
    ?year=2025&month=11&client_id=1&payment_status=unpaid&last_months=3
    """
    serializer = BillingReportFilterSerializer(data=request.query_params)
    serializer.is_valid(raise_exception=True)
    filters = serializer.validated_data
    
    queryset = BillingInvoice.objects.filter(
        status=BillingInvoice.InvoiceStatus.ISSUED
    )
    
    # Aplică filtre
    if filters.get('year'):
        queryset = queryset.filter(year=filters['year'])
    if filters.get('month'):
        queryset = queryset.filter(month=filters['month'])
    if filters.get('client_id'):
        queryset = queryset.filter(client_id=filters['client_id'])
    if filters.get('payment_status') and filters['payment_status'] != 'all':
        queryset = queryset.filter(payment_status=filters['payment_status'])
    if filters.get('last_months'):
        today = datetime.now()
        date_filters = Q()
        for i in range(filters['last_months']):
            target_date = today - timedelta(days=30 * i)
            date_filters |= Q(year=target_date.year, month=target_date.month)
        queryset = queryset.filter(date_filters)
    
    # Calculează sumarul
    totals = queryset.aggregate(
        total_subtotal=Sum('subtotal'),
        total_vat=Sum('vat_total'),
        total_amount=Sum('total'),
        total_paid=Sum('paid_amount'),
        total_due=Sum('due_amount')
    )
    
    # Breakdown pe status
    status_breakdown = {
        'paid': queryset.filter(payment_status='paid').count(),
        'partial': queryset.filter(payment_status='partial').count(),
        'unpaid': queryset.filter(payment_status='unpaid').count()
    }
    
    # Breakdown pe client
    by_client = []
    client_ids = queryset.values_list('client_id', flat=True).distinct()
    
    for client_id in client_ids:
        client_invoices = queryset.filter(client_id=client_id)
        client = Client.objects.get(id=client_id)
        client_totals = client_invoices.aggregate(
            subtotal=Sum('subtotal'),
            total=Sum('total'),
            paid=Sum('paid_amount'),
            due=Sum('due_amount')
        )
        by_client.append({
            'client_id': client_id,
            'client_name': client.denumire,
            'invoice_count': client_invoices.count(),
            'subtotal': float(client_totals['subtotal'] or 0),
            'total': float(client_totals['total'] or 0),
            'paid': float(client_totals['paid'] or 0),
            'due': float(client_totals['due'] or 0)
        })
    
    return Response({
        'invoice_count': queryset.count(),
        'totals': {
            'subtotal': float(totals['total_subtotal'] or 0),
            'vat': float(totals['total_vat'] or 0),
            'total': float(totals['total_amount'] or 0),
            'paid': float(totals['total_paid'] or 0),
            'due': float(totals['total_due'] or 0)
        },
        'status_breakdown': status_breakdown,
        'by_client': by_client
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def billing_export_excel(request):
    """Export raport facturare în Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    serializer = BillingReportFilterSerializer(data=request.query_params)
    serializer.is_valid(raise_exception=True)
    filters = serializer.validated_data
    
    queryset = BillingInvoice.objects.filter(
        status=BillingInvoice.InvoiceStatus.ISSUED
    ).select_related('client')
    
    # Aplică filtre (similar cu billing_report_summary)
    if filters.get('year'):
        queryset = queryset.filter(year=filters['year'])
    if filters.get('month'):
        queryset = queryset.filter(month=filters['month'])
    if filters.get('client_id'):
        queryset = queryset.filter(client_id=filters['client_id'])
    if filters.get('payment_status') and filters['payment_status'] != 'all':
        queryset = queryset.filter(payment_status=filters['payment_status'])
    
    # Creează workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raport Facturare"
    
    # Header
    headers = [
        'Nr.', 'Client', 'Serie/Număr', 'Data emiterii', 
        'Luna', 'An', 'Valoare fără TVA', 'TVA', 'Total',
        'Încasat', 'Sold', 'Status'
    ]
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Date
    for idx, invoice in enumerate(queryset.order_by('-year', '-month', '-issue_date'), 1):
        status_text = {
            'paid': 'Încasată',
            'partial': 'Parțial',
            'unpaid': 'Neîncasată'
        }.get(invoice.payment_status, invoice.payment_status)
        
        ws.append([
            idx,
            invoice.client.denumire,
            invoice.invoice_number_display,
            invoice.issue_date.strftime('%d.%m.%Y') if invoice.issue_date else '-',
            invoice.month,
            invoice.year,
            float(invoice.subtotal),
            float(invoice.vat_total),
            float(invoice.total),
            float(invoice.paid_amount),
            float(invoice.due_amount),
            status_text
        ])
    
    # Ajustează lățimea coloanelor
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
    
    # Salvează în buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="raport_facturare_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def billing_export_pdf(request):
    """Export raport facturare în PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return Response(
            {'detail': 'Biblioteca reportlab nu este instalată.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    serializer = BillingReportFilterSerializer(data=request.query_params)
    serializer.is_valid(raise_exception=True)
    filters = serializer.validated_data
    
    queryset = BillingInvoice.objects.filter(
        status=BillingInvoice.InvoiceStatus.ISSUED
    ).select_related('client')
    
    # Aplică filtre
    if filters.get('year'):
        queryset = queryset.filter(year=filters['year'])
    if filters.get('month'):
        queryset = queryset.filter(month=filters['month'])
    if filters.get('client_id'):
        queryset = queryset.filter(client_id=filters['client_id'])
    if filters.get('payment_status') and filters['payment_status'] != 'all':
        queryset = queryset.filter(payment_status=filters['payment_status'])
    
    # Creează PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm, leftMargin=1*cm,
        topMargin=1*cm, bottomMargin=1*cm
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1
    )
    
    elements = []
    elements.append(Paragraph("Raport Facturare", title_style))
    elements.append(Spacer(1, 20))
    
    # Tabel
    data = [[
        'Nr.', 'Client', 'Serie/Nr.', 'Data', 
        'Luna/An', 'Fără TVA', 'TVA', 'Total',
        'Încasat', 'Sold', 'Status'
    ]]
    
    for idx, invoice in enumerate(queryset.order_by('-year', '-month', '-issue_date'), 1):
        status_text = {
            'paid': 'Încasată',
            'partial': 'Parțial',
            'unpaid': 'Neîncasată'
        }.get(invoice.payment_status, invoice.payment_status)
        
        data.append([
            str(idx),
            invoice.client.denumire[:20],
            invoice.invoice_number_display,
            invoice.issue_date.strftime('%d.%m.%Y') if invoice.issue_date else '-',
            f'{invoice.month}/{invoice.year}',
            f'{invoice.subtotal:.2f}',
            f'{invoice.vat_total:.2f}',
            f'{invoice.total:.2f}',
            f'{invoice.paid_amount:.2f}',
            f'{invoice.due_amount:.2f}',
            status_text
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')])
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="raport_facturare_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response

